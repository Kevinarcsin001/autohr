"""PlatformImportAdapter 集成测试（任务 10）。

覆盖：
- detect_platform：3 个平台各一个 fixture Excel/ZIP；不支持格式
- import_package：Excel 结构化导入；附件包分流；不支持平台 422
- DB 副作用：Candidate/Source/Resume/ParsedStructure 写入；AsyncJob 入队
- 跨任务依赖：dedup_key 占位、async_jobs 占位入队
"""
from __future__ import annotations

import io
import uuid
import zipfile
from pathlib import Path

import pytest
from openpyxl import Workbook
from sqlalchemy import select, text

from app.core.db import AsyncSessionLocal
from app.models.async_job import AsyncJob
from app.models.candidate import (
    Candidate,
    CandidateResume,
    CandidateSource,
    ParsedStructure,
)
from app.models.team import Team
from app.models.user import User
from app.services.ingestion.platform_import import (
    PlatformImportAdapter,
    UnsupportedPlatformError,
)

# ============================================================================
# Fixtures
# ============================================================================


async def _purge_all() -> None:
    async with AsyncSessionLocal() as session:
        await session.execute(
            text(
                "TRUNCATE users, teams, team_invites, jobs, candidates, "
                "candidate_resumes, candidate_sources, parsed_structures, "
                "screening_results, scores, score_reasons, "
                "interview_questions, interview_feedbacks, dedup_matches, "
                "manual_overrides, llm_calls, async_jobs, audit_logs, "
                "email_configs, job_versions, job_hard_requirements "
                "RESTART IDENTITY CASCADE"
            )
        )
        await session.commit()


@pytest.fixture(autouse=True)
async def clean_db():
    await _purge_all()
    yield
    await _purge_all()


async def _make_team() -> uuid.UUID:
    async with AsyncSessionLocal() as session:
        team = Team(name=f"team-{uuid.uuid4().hex[:8]}")
        session.add(team)
        await session.commit()
        return team.id


# ============================================================================
# Fixture builders
# ============================================================================


def _boss_excel_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Boss直聘导出"
    ws.append(
        [
            "姓名",
            "应聘职位",
            "工作年限",
            "学历",
            "电话",
            "邮箱",
            "现居住地",
        ]
    )
    ws.append(
        ["张三", "高级前端工程师", 5, "本科", "13800138000", "zhang3@example.com", "上海"]
    )
    ws.append(
        ["李四", "Java 后端", 7, "硕士", "13900139000", "li4@example.com", "北京"]
    )
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _zhipin_excel_bytes() -> bytes:
    """注意：zhilian 关键词是 zhaopin/zhilian/智联。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "智联招聘"
    ws.append(["姓名", "性别", "年龄", "工作经验", "学历", "电话", "E-mail", "现居住地"])
    ws.append(["王五", "男", 28, 5, "本科", "13700137000", "wang5@example.com", "深圳"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _liepin_excel_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "猎聘导出"
    ws.append(["姓名", "当前公司", "当前职位", "工作年限", "学历", "电话", "邮箱"])
    ws.append(["赵六", "字节跳动", "技术专家", 8, "硕士", "13600136000", "zhao6@example.com"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _unknown_excel_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.append(["foo", "bar", "baz"])
    ws.append(["a", "b", "c"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _attachment_zip_bytes() -> bytes:
    """ZIP 含 2 个 PDF + 1 个 .txt（应跳过 txt）。"""
    pdf1 = (
        b"%PDF-1.4\n1 0 obj<</Type/Catalog/Pages 2 0 R>>endobj\n"
        b"xref\n0 4\n0000000000 65535 f \ntrailer<</Root 1 0 R>>\n%%EOF"
    )
    pdf2 = pdf1
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w") as zf:
        zf.writestr("boss_resume_1.pdf", pdf1)
        zf.writestr("boss_resume_2.pdf", pdf2)
        zf.writestr("readme.txt", b"ignore me")
    return buf.getvalue()


# ============================================================================
# detect_platform
# ============================================================================


async def test_detect_boss_excel() -> None:
    team_id = await _make_team()
    async with AsyncSessionLocal() as db:
        adapter = PlatformImportAdapter(db)
        result = await adapter.detect_platform(
            filename="boss_export.xlsx", content=_boss_excel_bytes()
        )
    assert result.platform == "boss"
    assert result.confidence >= 0.5
    assert result.package_kind == "excel"


async def test_detect_zhilian_excel() -> None:
    async with AsyncSessionLocal() as db:
        adapter = PlatformImportAdapter(db)
        result = await adapter.detect_platform(
            filename="zhaopin_export.xlsx",
            content=_zhipin_excel_bytes(),
        )
    assert result.platform == "zhilian"


async def test_detect_liepin_excel() -> None:
    async with AsyncSessionLocal() as db:
        adapter = PlatformImportAdapter(db)
        result = await adapter.detect_platform(
            filename="猎聘_2024.xlsx",
            content=_liepin_excel_bytes(),
        )
    assert result.platform == "liepin"


async def test_detect_unsupported_returns_none() -> None:
    async with AsyncSessionLocal() as db:
        adapter = PlatformImportAdapter(db)
        result = await adapter.detect_platform(
            filename="random.xlsx",
            content=_unknown_excel_bytes(),
        )
    assert result.platform is None
    assert result.confidence < result.threshold


async def test_detect_attachment_zip() -> None:
    async with AsyncSessionLocal() as db:
        adapter = PlatformImportAdapter(db)
        result = await adapter.detect_platform(
            filename="boss_resumes.zip",
            content=_attachment_zip_bytes(),
        )
    # ZIP 成员名含 boss → confidence ≥ threshold
    assert result.platform == "boss"
    assert result.package_kind == "attachment_zip"


# ============================================================================
# import_package：Excel 结构化
# ============================================================================


async def test_import_boss_excel_writes_db() -> None:
    team_id = await _make_team()
    async with AsyncSessionLocal() as db:
        adapter = PlatformImportAdapter(db)
        result = await adapter.import_package(
            team_id=team_id,
            filename="boss.xlsx",
            content=_boss_excel_bytes(),
        )
        await db.commit()

    assert result.platform == "boss"
    assert result.imported == 2
    assert len(result.candidates) == 2
    names = {c.name for c in result.candidates}
    assert names == {"张三", "李四"}

    # DB 副作用：2 Candidate + 2 Source + 2 Resume + 2 ParsedStructure
    async with AsyncSessionLocal() as session:
        cands = (await session.execute(select(Candidate))).scalars().all()
        srcs = (await session.execute(select(CandidateSource))).scalars().all()
        resumes = (await session.execute(select(CandidateResume))).scalars().all()
        structs = (await session.execute(select(ParsedStructure))).scalars().all()
    assert len(cands) == 2
    assert len(srcs) == 2
    assert len(resumes) == 2
    assert len(structs) == 2
    # 结构化路径不写 AsyncJob（不走 parser）
    async with AsyncSessionLocal() as session:
        jobs = (await session.execute(select(AsyncJob))).scalars().all()
    assert len(jobs) == 0
    # parse_status = success（跳过 parser）
    assert all(r.parse_status == "success" for r in resumes)
    # ParsedStructure.data 含归一化字段
    assert "education" in structs[0].data
    assert structs[0].data["education"] in {"bachelor", "master"}


async def test_import_skips_row_without_identity() -> None:
    """Excel 行缺 phone+email → rejected(missing_identity)。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Boss直聘"
    ws.append(["姓名", "电话", "邮箱"])
    ws.append(["无名氏", None, None])  # missing identity
    ws.append(["张三", "13800138000", "z@x.com"])  # ok
    buf = io.BytesIO()
    wb.save(buf)

    team_id = await _make_team()
    async with AsyncSessionLocal() as db:
        adapter = PlatformImportAdapter(db)
        result = await adapter.import_package(
            team_id=team_id,
            filename="boss.xlsx",
            content=buf.getvalue(),
        )
        await db.commit()
    assert result.imported == 1
    assert result.rejected == 1
    rejected = next(c for c in result.candidates if c.status == "rejected")
    assert rejected.reject_reason == "missing_identity"


async def test_import_dedup_by_identity() -> None:
    """同一 phone/email 二次导入 → rejected(duplicate)。"""
    team_id = await _make_team()
    async with AsyncSessionLocal() as db:
        adapter = PlatformImportAdapter(db)
        await adapter.import_package(
            team_id=team_id, filename="boss.xlsx", content=_boss_excel_bytes()
        )
        await db.commit()

    async with AsyncSessionLocal() as db:
        adapter = PlatformImportAdapter(db)
        result = await adapter.import_package(
            team_id=team_id, filename="boss.xlsx", content=_boss_excel_bytes()
        )
        await db.commit()
    assert result.imported == 0
    assert result.rejected == 2
    assert all(c.reject_reason == "duplicate" for c in result.candidates)


# ============================================================================
# import_package：附件包分流
# ============================================================================


async def test_import_attachment_zip_writes_db_and_enqueues() -> None:
    team_id = await _make_team()
    async with AsyncSessionLocal() as db:
        adapter = PlatformImportAdapter(db)
        result = await adapter.import_package(
            team_id=team_id,
            filename="boss_resumes.zip",
            content=_attachment_zip_bytes(),
        )
        await db.commit()

    assert result.platform == "boss"
    assert result.package_kind == "attachment_zip"
    assert result.imported == 2  # 2 PDF；readme.txt 跳过
    assert len(result.candidates) == 2

    async with AsyncSessionLocal() as session:
        resumes = (await session.execute(select(CandidateResume))).scalars().all()
        jobs = (
            await session.execute(select(AsyncJob).where(AsyncJob.task_type == "parse"))
        ).scalars().all()
    assert len(resumes) == 2
    assert len(jobs) == 2
    assert all(r.parse_status == "pending" for r in resumes)
    assert all(j.status == "queued" for j in jobs)


async def test_import_attachment_zip_mime_mismatch_rejected() -> None:
    """附件扩展名 .pdf 但内容是 PNG → 该附件 rejected。"""
    png = (
        b"\x89PNG\r\n\x1a\n"
        b"\x00\x00\x00\rIHDR\x00\x00\x00\x01\x00\x00\x00\x01\x08\x06"
        b"\x00\x00\x00\x1f\x15\xc4\x89"
        b"\x00\x00\x00\x00IEND\xaeB`\x82"
    )
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w") as zf:
        zf.writestr("boss_resume.pdf", png)
    team_id = await _make_team()
    async with AsyncSessionLocal() as db:
        adapter = PlatformImportAdapter(db)
        result = await adapter.import_package(
            team_id=team_id,
            filename="boss.zip",
            content=buf.getvalue(),
        )
        await db.commit()
    assert result.imported == 0
    assert result.rejected == 1
    assert result.candidates[0].reject_reason == "invalid_structure"


# ============================================================================
# 不支持平台 → 422
# ============================================================================


async def test_import_unsupported_raises() -> None:
    team_id = await _make_team()
    async with AsyncSessionLocal() as db:
        adapter = PlatformImportAdapter(db)
        with pytest.raises(UnsupportedPlatformError) as exc_info:
            await adapter.import_package(
                team_id=team_id,
                filename="random.xlsx",
                content=_unknown_excel_bytes(),
            )
    err = exc_info.value
    assert err.code == "unsupported_platform"
    assert err.detection.platform is None
    assert err.support_feedback_url.startswith("http")
