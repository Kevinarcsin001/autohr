"""ExportService 集成测试（任务 22）。

策略：
1. ``_FakeStorage`` 替代真 MinIO；记录 put / signed_url / exists 调用
2. ``request_export``：小数据集走 sync，立即返回 download_url；大数据集入 async_jobs
3. ``_generate`` 写 xlsx 后 ``signed_url`` 有 5min expires
4. ``get_signed_download_url``：file_key 前缀校验 + 对象必须存在
5. Excel 内容：列顺序固定、手机号脱敏、空数据也合法
6. ``_make_file_key`` / ``_validate_file_key_prefix`` / ``_mask_phone`` 单元
"""
from __future__ import annotations

import io
import uuid
from typing import Any

import pytest
from openpyxl import load_workbook
from sqlalchemy import select, text

from app.core.db import AsyncSessionLocal
from app.core.middleware.error_handler import NotFoundError
from app.models.async_job import AsyncJob
from app.models.candidate import (
    Candidate,
    CandidateResume,
    CandidateSource,
    ParsedStructure,
)
from app.models.interview import InterviewQuestion
from app.models.job import Job
from app.models.score import Score, ScoreReason
from app.models.screening import ScreeningResult
from app.models.team import Team
from app.models.user import User
from app.services.export import (
    EXPORT_ASYNC_THRESHOLD,
    EXPORT_COLUMNS,
    ExportService,
    _make_file_key,
    _mask_phone,
    _validate_file_key_prefix,
)


# ============================================================================
# DB 清理
# ============================================================================


async def _purge_db() -> None:
    async with AsyncSessionLocal() as session:
        await session.execute(
            text(
                "TRUNCATE users, teams, team_invites, jobs, candidates, "
                "candidate_resumes, candidate_sources, parsed_structures, "
                "screening_results, scores, score_reasons, "
                "interview_questions, interview_feedbacks, dedup_matches, "
                "manual_overrides, llm_calls, async_jobs, audit_logs, "
                "email_configs, job_versions, job_hard_requirements "
                "RESTART IDENTITY CASCADE"
            )
        )
        await session.commit()


@pytest.fixture(autouse=True)
async def clean_db() -> None:
    await _purge_db()
    yield
    await _purge_db()


# ============================================================================
# Fake Storage
# ============================================================================


class _FakeStorage:
    """记录所有调用 + 内存存储 + signed_url 包含 expires 参数。"""

    def __init__(self) -> None:
        self._store: dict[str, bytes] = {}
        self.put_calls: list[tuple[str, bytes, str, bool]] = []
        self.signed_url_calls: list[tuple[str, int | None, str]] = []
        self.exists_calls: list[str] = []

    async def put(
        self, key: str, data: bytes, *, mime: str, encrypt: bool = True
    ) -> None:
        self.put_calls.append((key, data, mime, encrypt))
        self._store[key] = data

    async def get(self, key: str) -> bytes:
        return self._store[key]

    async def exists(self, key: str) -> bool:
        self.exists_calls.append(key)
        return key in self._store

    async def delete(self, key: str) -> None:
        self._store.pop(key, None)

    async def signed_url(
        self, key: str, *, expires: int | None = None, method: str = "GET"
    ) -> str:
        self.signed_url_calls.append((key, expires, method))
        return f"https://fake.local/{key}?expires={expires}&method={method}"


# ============================================================================
# Seed 工具
# ============================================================================


async def _seed_full_setup(
    *, n_candidates: int = 3, with_score: bool = True, with_interview: bool = True
) -> tuple[Team, User, Job, list[Candidate]]:
    """team + user + job + N candidates（含 source/resume/structure/score/screen/interview）。"""
    async with AsyncSessionLocal() as session:
        team = Team(name=f"team-{uuid.uuid4().hex[:8]}")
        session.add(team)
        await session.flush()

        user = User(
            email=f"u-{uuid.uuid4().hex[:8]}@x.com",
            password_hash="x",
            name="hr",
            team_id=team.id,
        )
        session.add(user)
        await session.flush()

        job = Job(
            team_id=team.id,
            title="Eng",
            jd_text="Python 工程师",
            status="active",
            created_by=user.id,
        )
        session.add(job)
        await session.flush()

        candidates: list[Candidate] = []
        for i in range(n_candidates):
            cand = Candidate(
                team_id=team.id,
                dedup_key=f"test:{uuid.uuid4()}",
                name=f"cand-{i}",
                phone=f"1380000000{i}",
                email=f"c{i}@x.com",
            )
            session.add(cand)
            await session.flush()

            src = CandidateSource(
                candidate_id=cand.id, source_type="upload"
            )
            session.add(src)
            await session.flush()

            resume = CandidateResume(
                candidate_id=cand.id,
                source_id=src.id,
                file_storage_key=f"k-{i}",
                file_mime="application/pdf",
                parse_status="success",
                parsed_text="Python 5年",
            )
            session.add(resume)
            await session.flush()

            session.add(ParsedStructure(
                resume_id=resume.id,
                data={
                    "structure": {
                        "name": f"cand-{i}",
                        "phone": f"1380000000{i}",
                        "email": f"c{i}@x.com",
                        "education": "master",
                        "years_of_experience": 5,
                        "current_company": "ACME",
                        "skills": ["Python", "FastAPI"],
                    },
                    "status": "extracted",
                },
            ))

            session.add(ScreeningResult(
                job_id=job.id,
                candidate_id=cand.id,
                disqualified=(i == 0),
                reasons=["学历不达标"] if i == 0 else None,
            ))

            if with_score:
                score = Score(
                    job_id=job.id,
                    candidate_id=cand.id,
                    total=80 + i,
                    skill=85, experience=80, education=75,
                    stability=80, potential=85,
                )
                session.add(score)
                await session.flush()
                session.add(ScoreReason(
                    score_id=score.id,
                    type="recommend",
                    bullet_points=["Python 技能匹配", "5 年经验"],
                ))

            if with_interview:
                session.add(InterviewQuestion(
                    candidate_id=cand.id,
                    job_id=job.id,
                    batch_id=uuid.uuid4(),
                    dimension="skill",
                    question="聊聊 Python 项目",
                    sort_order=0,
                ))

            candidates.append(cand)

        await session.commit()
        return team, user, job, candidates


# ============================================================================
# 单元：纯函数
# ============================================================================


class TestHelpers:
    def test_make_file_key_format(self) -> None:
        team_id = uuid.uuid4()
        job_id = uuid.uuid4()
        key = _make_file_key(team_id, job_id)
        assert key.startswith(f"exports/{team_id}/{job_id}/")
        assert key.endswith(".xlsx")

    def test_validate_file_key_prefix_ok(self) -> None:
        team_id = uuid.uuid4()
        job_id = uuid.uuid4()
        key = f"exports/{team_id}/{job_id}/abc.xlsx"
        _validate_file_key_prefix(key, team_id)  # not raise

    def test_validate_file_key_prefix_wrong_team(self) -> None:
        team_a = uuid.uuid4()
        team_b = uuid.uuid4()
        key = f"exports/{team_a}/{uuid.uuid4()}/abc.xlsx"
        with pytest.raises(NotFoundError):
            _validate_file_key_prefix(key, team_b)

    def test_validate_file_key_prefix_malformed(self) -> None:
        bad_team = uuid.uuid4()
        with pytest.raises(NotFoundError):
            _validate_file_key_prefix("random/abc.xlsx", bad_team)

    def test_mask_phone_normal(self) -> None:
        assert _mask_phone("13812345678") == "138****5678"

    def test_mask_phone_short(self) -> None:
        assert _mask_phone("12345") == "12345"

    def test_mask_phone_none(self) -> None:
        assert _mask_phone(None) == ""


# ============================================================================
# request_export 同步 vs 异步
# ============================================================================


class TestRequestExport:
    async def test_sync_small_dataset(self) -> None:
        team, user, job, cands = await _seed_full_setup(n_candidates=3)
        storage = _FakeStorage()
        async with AsyncSessionLocal() as session:
            service = ExportService(session, storage=storage)
            result = await service.request_export(
                team_id=team.id,
                user_id=user.id,
                job_id=job.id,
                filters=None,
                format="xlsx",
            )
            await session.commit()

        assert result["mode"] == "sync"
        assert "download_url" in result
        assert result["expires_in"] == 300
        assert result["row_count"] == 3
        assert result["file_key"].startswith(f"exports/{team.id}/{job.id}/")
        assert result["file_size"] > 0
        # storage 被调用 put + signed_url
        assert len(storage.put_calls) == 1
        assert len(storage.signed_url_calls) == 1
        # signed_url 走 GET，5min 过期
        _, expires, method = storage.signed_url_calls[0]
        assert expires == 300
        assert method == "GET"

    async def test_async_large_dataset(self) -> None:
        team, user, job, _ = await _seed_full_setup(n_candidates=3)
        storage = _FakeStorage()

        # monkeypatch 阈值改为 2 → 3 行触发异步入队
        import app.services.export as export_mod
        original = export_mod.EXPORT_ASYNC_THRESHOLD
        export_mod.EXPORT_ASYNC_THRESHOLD = 2
        try:
            async with AsyncSessionLocal() as session:
                service = ExportService(session, storage=storage)
                result = await service.request_export(
                    team_id=team.id,
                    user_id=user.id,
                    job_id=job.id,
                    filters=None,
                    format="xlsx",
                )
                await session.commit()
        finally:
            export_mod.EXPORT_ASYNC_THRESHOLD = original

        assert result["mode"] == "async"
        assert "job_id" in result
        assert result["row_count"] == 3
        # storage 不应被调用（异步走 celery 后才生成）
        assert len(storage.put_calls) == 0

        # async_jobs 应有一行 task_type=export
        async with AsyncSessionLocal() as session:
            jobs = (
                await session.execute(
                    select(AsyncJob).where(AsyncJob.task_type == "export")
                )
            ).scalars().all()
            assert len(jobs) == 1
            assert jobs[0].status == "queued"
            assert str(jobs[0].target_id) == str(user.id)
            assert jobs[0].payload["job_id"] == str(job.id)
            assert jobs[0].payload["team_id"] == str(team.id)
            assert jobs[0].payload["format"] == "xlsx"

    async def test_request_export_cross_team_job_404(self) -> None:
        """跨 team 访问 job → NotFoundError（不暴露存在性）。"""
        team, user, job, _ = await _seed_full_setup(n_candidates=1)
        other_team = uuid.uuid4()
        storage = _FakeStorage()
        async with AsyncSessionLocal() as session:
            service = ExportService(session, storage=storage)
            with pytest.raises(NotFoundError):
                await service.request_export(
                    team_id=other_team,
                    user_id=user.id,
                    job_id=job.id,
                )


# ============================================================================
# get_signed_download_url
# ============================================================================


class TestSignedDownloadUrl:
    async def test_ok(self) -> None:
        team, user, job, _ = await _seed_full_setup(n_candidates=1)
        storage = _FakeStorage()
        async with AsyncSessionLocal() as session:
            service = ExportService(session, storage=storage)
            file_key, _ = await service._generate(
                job_id=job.id, team_id=team.id, filters={},
            )
            url = await service.get_signed_download_url(
                team_id=team.id, file_key=file_key,
            )
            await session.commit()
        assert url.startswith("https://fake.local/")
        assert "expires=300" in url

    async def test_cross_team_404(self) -> None:
        team, _, job, _ = await _seed_full_setup(n_candidates=1)
        storage = _FakeStorage()
        other_team = uuid.uuid4()
        async with AsyncSessionLocal() as session:
            service = ExportService(session, storage=storage)
            file_key, _ = await service._generate(
                job_id=job.id, team_id=team.id, filters={},
            )
            with pytest.raises(NotFoundError):
                await service.get_signed_download_url(
                    team_id=other_team, file_key=file_key,
                )

    async def test_object_missing_404(self) -> None:
        """file_key 前缀 OK 但对象不存在 → 404。"""
        team = uuid.uuid4()
        job = uuid.uuid4()
        file_key = f"exports/{team}/{job}/missing.xlsx"
        storage = _FakeStorage()
        async with AsyncSessionLocal() as session:
            service = ExportService(session, storage=storage)
            with pytest.raises(NotFoundError):
                await service.get_signed_download_url(
                    team_id=team, file_key=file_key,
                )


# ============================================================================
# Excel 内容
# ============================================================================


class TestExcelContent:
    async def test_columns_match_schema(self) -> None:
        """Excel 第一行表头 = EXPORT_COLUMNS。"""
        team, user, job, _ = await _seed_full_setup(n_candidates=2)
        storage = _FakeStorage()
        async with AsyncSessionLocal() as session:
            service = ExportService(session, storage=storage)
            file_key, _ = await service._generate(
                job_id=job.id, team_id=team.id, filters={},
            )
            await session.commit()

        data = storage._store[file_key]
        wb = load_workbook(io.BytesIO(data), read_only=True)
        ws = wb.active
        header = [c.value for c in next(ws.iter_rows(max_row=1))]
        assert tuple(header) == EXPORT_COLUMNS

    async def test_phone_masked_in_excel(self) -> None:
        """导出 Excel 中手机号必须是脱敏格式。"""
        team, user, job, _ = await _seed_full_setup(n_candidates=1)
        storage = _FakeStorage()
        async with AsyncSessionLocal() as session:
            service = ExportService(session, storage=storage)
            file_key, _ = await service._generate(
                job_id=job.id, team_id=team.id, filters={},
            )
            await session.commit()

        data = storage._store[file_key]
        wb = load_workbook(io.BytesIO(data), read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        # 第 2 行第 1 列是姓名；第 3 列是电话
        phone_cell = rows[1][2]
        assert "****" in str(phone_cell)
        assert phone_cell == "138****0000"

    async def test_empty_dataset_still_valid(self) -> None:
        """无候选人时也能生成合法 xlsx（仅表头）。"""
        async with AsyncSessionLocal() as session:
            team = Team(name=f"empty-{uuid.uuid4().hex[:6]}")
            session.add(team)
            await session.flush()
            user = User(
                email="empty@x.com", password_hash="x",
                name="hr", team_id=team.id,
            )
            session.add(user)
            await session.flush()
            job = Job(
                team_id=team.id, title="x", jd_text="x",
                status="active", created_by=user.id,
            )
            session.add(job)
            await session.flush()

            storage = _FakeStorage()
            service = ExportService(session, storage=storage)
            file_key, size = await service._generate(
                job_id=job.id, team_id=team.id, filters={},
            )
            await session.commit()

        assert size > 0
        data = storage._store[file_key]
        wb = load_workbook(io.BytesIO(data), read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        # 仅表头 1 行
        assert len(rows) == 1


# ============================================================================
# run_export_handler（worker 层集成）
# ============================================================================


class TestRunExportHandler:
    async def test_handler_writes_file_and_returns_summary(self) -> None:
        """run_export_handler 直接调 _generate + _notify_user，返回 dict。"""
        from app.workers.tasks import run_export_handler

        team, user, job, _ = await _seed_full_setup(n_candidates=2)

        # 用 monkeypatch 替换 ExportService 内部 storage 单例
        import app.services.export as export_mod
        fake_storage = _FakeStorage()
        original_init = export_mod.ExportService.__init__

        def _patched_init(self, db, *, storage=None):
            original_init(self, db, storage=fake_storage)

        export_mod.ExportService.__init__ = _patched_init
        try:
            result = await run_export_handler(
                target_id=user.id,
                payload={
                    "job_id": str(job.id),
                    "team_id": str(team.id),
                    "user_id": str(user.id),
                    "format": "xlsx",
                    "filters": {},
                },
            )
        finally:
            export_mod.ExportService.__init__ = original_init

        assert result is not None
        assert result["file_key"].startswith(f"exports/{team.id}/{job.id}/")
        assert result["row_count"] == 2
        assert result["email_sent"] is True
        assert len(fake_storage.put_calls) == 1

    async def test_handler_missing_job_id_raises(self) -> None:
        from app.workers.tasks import run_export_handler

        with pytest.raises(ValueError, match="job_id"):
            await run_export_handler(target_id=None, payload={})
