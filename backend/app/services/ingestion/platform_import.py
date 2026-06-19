"""PlatformImportAdapter（任务 10）：招聘平台导出包识别 + 导入。

流程：
1. ``detect_platform(filename, content)`` 评分（文件名 + Excel 表头 / ZIP 成员）→
   取最高分；若 < 阈值 → 返回 platform=None。
2. ``import_package(filename, content, team_id)``：
   - ZIP → 检查内容；若为 Excel/JSON → 结构化路径；若全为简历附件 → 走任务 9 路径
   - 单 Excel → 结构化路径：mapper 把每行映射成 CandidateStructure → 校验 → 写库
   - 其他 → 422 UnsupportedPlatformError

跨任务依赖：
- 任务 9 file_upload：附件包分流时复用其 storage + DB 写入逻辑
- 任务 14 dedup：本任务用 ``platform:{platform}:{normalized_identity}`` 作为占位 dedup_key
- 任务 13 parser：附件包入队后由 task 12 消费触发

事务策略：service 接收 session，不 commit。
"""
from __future__ import annotations

import io
import re
import uuid
import zipfile
from typing import Any

import magic
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.adapters.storage import S3StorageAdapter, get_storage
from app.core.config import settings
from app.core.logging import get_logger
from app.core.middleware.error_handler import ValidationError
from app.models.async_job import AsyncJob
from app.models.candidate import (
    Candidate,
    CandidateResume,
    CandidateSource,
)
from app.schemas.platform import (
    CandidateStructure,
    DetectionResult,
    DetectionSignal,
    ImportedCandidateItem,
    Platform,
    PlatformImportResult,
    PlatformPackageKind,
)

logger = get_logger(__name__)


# ============================================================================
# 异常
# ============================================================================


class UnsupportedPlatformError(ValidationError):
    """不支持的平台格式（422 + 反馈入口）。"""

    def __init__(
        self,
        message: str,
        *,
        detection: DetectionResult,
        support_feedback_url: str,
    ) -> None:
        super().__init__(
            message,
            code="unsupported_platform",
            detection=detection.model_dump(mode="json"),
            support_feedback_url=support_feedback_url,
        )
        self.detection = detection
        self.support_feedback_url = support_feedback_url


# ============================================================================
# 平台特征库（识别 + 字段映射）
# ============================================================================


# 文件名关键词 → 平台（小写匹配）
_FILENAME_KEYWORDS: dict[Platform, list[str]] = {
    "boss": ["boss"],
    "zhilian": ["zhaopin", "zhilian", "智联"],
    "liepin": ["liepin", "猎聘"],
}

# Excel 表头关键词 → 平台
_HEADER_KEYWORDS: dict[Platform, list[str]] = {
    "boss": ["boss", "boss直聘", "boss 直聘"],
    "zhilian": ["智联", "zhaopin", "zhaopin.com"],
    "liepin": ["猎聘", "liepin", "liepin.com"],
}

# 平台 → 列名映射（CandidateStructure 字段）
# 列名小写 + 去空白匹配
_COLUMN_MAP: dict[Platform, dict[str, str]] = {
    "boss": {
        "姓名": "name",
        "电话": "phone",
        "邮箱": "email",
        "手机": "phone",
        "性别": "gender",
        "年龄": "age",
        "学历": "education",
        "工作年限": "years_experience",
        "工作经验": "years_experience",
        "应聘职位": "applied_position",
        "应聘岗位": "applied_position",
        "当前公司": "current_company",
        "当前职位": "current_title",
        "现居地": "location",
        "现居住地": "location",
        "城市": "location",
    },
    "zhilian": {
        "姓名": "name",
        "电话": "phone",
        "手机号": "phone",
        "e-mail": "email",
        "email": "email",
        "邮箱": "email",
        "性别": "gender",
        "年龄": "age",
        "学历": "education",
        "最高学历": "education",
        "工作经验": "years_experience",
        "工作年限": "years_experience",
        "应聘职位": "applied_position",
        "求职意向": "applied_position",
        "现居住地": "location",
        "所在地区": "location",
    },
    "liepin": {
        "姓名": "name",
        "电话": "phone",
        "手机": "phone",
        "邮箱": "email",
        "e-mail": "email",
        "性别": "gender",
        "年龄": "age",
        "学历": "education",
        "工作年限": "years_experience",
        "当前公司": "current_company",
        "当前职位": "current_title",
        "职位": "current_title",
        "行业": "applied_position",  # 猎聘常列"求职行业"
        "现居城市": "location",
        "所在地": "location",
    },
}

# 教育 / 性别归一化
_EDU_MAP = {
    "高中": "high_school",
    "中专": "high_school",
    "大专": "high_school",
    "专科": "high_school",
    "本科": "bachelor",
    "学士": "bachelor",
    "硕士": "master",
    "研究生": "master",
    "博士": "phd",
    "mba": "master",
    "high school": "high_school",
    "bachelor": "bachelor",
    "master": "master",
    "phd": "phd",
    "doctor": "phd",
}
_GENDER_MAP = {
    "男": "male",
    "女": "female",
    "male": "male",
    "female": "female",
    "m": "male",
    "f": "female",
}

_ATTACHMENT_EXT = {".pdf", ".doc", ".docx", ".png", ".jpg", ".jpeg"}
_PLATFORM_ALLOWED_MIME = {
    "application/pdf",
    "application/msword",
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    "image/png",
    "image/jpeg",
}


# ============================================================================
# 工具
# ============================================================================


def _normalize_text(s: str) -> str:
    return re.sub(r"\s+", "", s).lower()


def _is_zip_bytes(content: bytes) -> bool:
    return content[:4] == b"PK\x03\x04" or content[:4] == b"PK\x05\x06"


def _is_excel_bytes(content: bytes) -> bool:
    # XLSX (zip-based) 或 XLS (CFB)
    return (content[:4] == b"PK\x03\x04" and content[
        4:8
    ] != b"") or content[:8] == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"


def _sniff_mime(content: bytes) -> str:
    try:
        return magic.from_buffer(content, mime=True)
    except Exception:  # noqa: BLE001
        return "application/octet-stream"


def _normalize_identity(c: dict[str, Any]) -> str | None:
    """构造稳定的身份串（用于 dedup_key 占位）。

    TODO(task-14): 真实 dedup_key 由 dedup service 接管（normalize(name)+last4(phone)+prefix(email)）。
    """
    name = (c.get("name") or "").strip()
    phone = (c.get("phone") or "").strip()
    email = (c.get("email") or "").strip().lower()
    if not name and not phone and not email:
        return None
    return f"{name}|{phone[-4:] if phone else ''}|{email.split('@')[0][:8] if email else ''}"


# ============================================================================
# 评分（detect_platform）
# ============================================================================


def _score_filename(filename: str) -> dict[Platform, list[DetectionSignal]]:
    name_norm = _normalize_text(filename)
    out: dict[Platform, list[DetectionSignal]] = {p: [] for p in _FILENAME_KEYWORDS}
    for platform, kws in _FILENAME_KEYWORDS.items():
        for kw in kws:
            if _normalize_text(kw) in name_norm:
                out[platform].append(
                    DetectionSignal(
                        source="filename",
                        weight=0.3,
                        matched=kw,
                    )
                )
    return out


def _score_excel_header(content: bytes) -> tuple[
    dict[Platform, list[DetectionSignal]], list[str] | None
]:
    """读 Excel sheet title + 表头，返回每个平台的匹配信号。"""
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception:  # noqa: BLE001
        return {p: [] for p in _HEADER_KEYWORDS}, None
    if not wb.sheetnames:
        return {p: [] for p in _HEADER_KEYWORDS}, None
    sheet_title = wb.sheetnames[0] or ""
    ws = wb[sheet_title]
    headers: list[str] = []
    for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
        non_empty = [str(c).strip() for c in (row or []) if c]
        if len(non_empty) >= 3:
            headers = non_empty
            break
    wb.close()

    # 评分基线：sheet title（高权重，平台导出通常 sheet 名即平台名）
    title_blob = _normalize_text(sheet_title)
    headers_blob = _normalize_text("|".join(headers)) if headers else ""
    out: dict[Platform, list[DetectionSignal]] = {p: [] for p in _HEADER_KEYWORDS}
    for platform, kws in _HEADER_KEYWORDS.items():
        for kw in kws:
            kw_norm = _normalize_text(kw)
            if kw_norm in title_blob:
                out[platform].append(
                    DetectionSignal(
                        source="header",
                        weight=0.5,
                        matched=f"sheet:{kw}",
                    )
                )
            elif kw_norm in headers_blob:
                out[platform].append(
                    DetectionSignal(
                        source="header",
                        weight=0.4,
                        matched=f"header:{kw}",
                    )
                )
    return out, headers or None


def _score_zip_members(
    content: bytes,
) -> tuple[dict[Platform, list[DetectionSignal]], list[str]]:
    """读 ZIP 成员文件名，识别平台。"""
    out: dict[Platform, list[DetectionSignal]] = {p: [] for p in _FILENAME_KEYWORDS}
    members: list[str] = []
    try:
        with zipfile.ZipFile(io.BytesIO(content)) as zf:
            members = [n for n in zf.namelist() if not n.endswith("/")]
            blob = _normalize_text("|".join(members))
    except zipfile.BadZipFile:
        return out, []
    for platform, kws in _FILENAME_KEYWORDS.items():
        for kw in kws:
            if _normalize_text(kw) in blob:
                out[platform].append(
                    DetectionSignal(
                        source="zip_member",
                        weight=0.2,
                        matched=kw,
                    )
                )
    return out, members


def _merge_signals(
    *signal_maps: dict[Platform, list[DetectionSignal]],
) -> dict[Platform, list[DetectionSignal]]:
    merged: dict[Platform, list[DetectionSignal]] = {p: [] for p in _FILENAME_KEYWORDS}
    for sm in signal_maps:
        for p, sigs in sm.items():
            merged[p].extend(sigs)
    return merged


def _score_platform(signals: list[DetectionSignal]) -> float:
    """单平台得分 = min(sum(weights), 1.0)。"""
    return min(sum(s.weight for s in signals), 1.0)


# ============================================================================
# PlatformImportAdapter
# ============================================================================


class PlatformImportAdapter:
    """招聘平台导出包导入适配器。"""

    def __init__(
        self,
        db: AsyncSession,
        storage: S3StorageAdapter | None = None,
    ) -> None:
        self.db = db
        self.storage = storage or get_storage()
        self.threshold = settings.PLATFORM_DETECT_MIN_CONFIDENCE
        self.feedback_url = settings.PLATFORM_SUPPORT_FEEDBACK_URL

    # ----- 阶段 1：detect_platform -----

    async def detect_platform(
        self,
        *,
        filename: str,
        content: bytes,
    ) -> DetectionResult:
        """评分检测平台类型 + 包类型。"""
        fname_signals = _score_filename(filename)

        # 包类型 + 内部信号
        package_kind: PlatformPackageKind | None = None
        internal_signals: dict[Platform, list[DetectionSignal]] = {
            p: [] for p in _FILENAME_KEYWORDS
        }
        if _is_zip_bytes(content):
            # 先尝试当 Excel（xlsx 也是 zip）
            try:
                wb = load_workbook(io.BytesIO(content), read_only=True)
                is_excel = True
                wb.close()
            except Exception:  # noqa: BLE001
                is_excel = False

            if is_excel:
                package_kind = "excel"
                internal_signals, _ = _score_excel_header(content)
            else:
                package_kind = "attachment_zip"
                internal_signals, _ = _score_zip_members(content)
        else:
            # 单文件：尝试 Excel
            try:
                wb = load_workbook(io.BytesIO(content), read_only=True)
                is_excel = True
                wb.close()
            except Exception:  # noqa: BLE001
                is_excel = False
            if is_excel:
                package_kind = "excel"
                internal_signals, _ = _score_excel_header(content)

        signals = _merge_signals(fname_signals, internal_signals)
        scores: dict[Platform, float] = {
            p: _score_platform(sigs) for p, sigs in signals.items()
        }

        # 选最高分平台
        best_platform: Platform | None = None
        best_score = 0.0
        for p, sc in scores.items():
            if sc > best_score:
                best_score = sc
                best_platform = p
        if best_score < self.threshold:
            best_platform = None

        return DetectionResult(
            platform=best_platform,
            confidence=best_score,
            package_kind=package_kind,
            threshold=self.threshold,
            signals=[
                s
                for sigs in signals.values()
                for s in sigs
            ],
            scores=scores,
        )

    # ----- 阶段 2：import_package -----

    async def import_package(
        self,
        *,
        team_id: uuid.UUID,
        filename: str,
        content: bytes,
    ) -> PlatformImportResult:
        """主入口：先 detect，再按 package_kind 分流。"""
        detection = await self.detect_platform(
            filename=filename, content=content
        )
        if detection.platform is None or detection.package_kind is None:
            raise UnsupportedPlatformError(
                "暂不支持的平台格式",
                detection=detection,
                support_feedback_url=self.feedback_url,
            )

        if detection.package_kind == "excel":
            return await self._import_excel(
                team_id=team_id,
                platform=detection.platform,
                content=content,
            )
        else:
            return await self._import_attachment_zip(
                team_id=team_id,
                platform=detection.platform,
                content=content,
            )

    # ----- Excel mapper -----

    async def _import_excel(
        self,
        *,
        team_id: uuid.UUID,
        platform: Platform,
        content: bytes,
    ) -> PlatformImportResult:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            wb.close()
            return PlatformImportResult(
                platform=platform,
                package_kind="excel",
                candidates=[],
                imported=0,
                rejected=0,
            )
        headers = [str(c or "").strip() for c in header_row]
        col_map = _COLUMN_MAP[platform]
        # 列名 → 字段名（取第一次出现）
        header_to_field: dict[int, str] = {}
        for i, h in enumerate(headers):
            if not h:
                continue
            for src, dst in col_map.items():
                if _normalize_text(src) == _normalize_text(h):
                    header_to_field[i] = dst
                    break

        results: list[ImportedCandidateItem] = []
        for raw_row in rows_iter:
            if raw_row is None:
                continue
            # 跳过空行
            if all(c is None or str(c).strip() == "" for c in raw_row):
                continue
            mapped: dict[str, Any] = {}
            raw: dict[str, str] = {}
            for i, cell in enumerate(raw_row):
                if i >= len(headers):
                    break
                h = headers[i]
                val = "" if cell is None else str(cell).strip()
                if h and val:
                    raw[h] = val
                field = header_to_field.get(i)
                if field and val:
                    if field in mapped:
                        # 多列映射到同字段时取非空（如 phone / 手机）
                        continue
                    mapped[field] = val
            item = await self._persist_one(
                team_id=team_id,
                platform=platform,
                mapped=mapped,
                raw=raw,
            )
            results.append(item)
        wb.close()

        imported = sum(1 for r in results if r.status == "ok")
        rejected = len(results) - imported
        return PlatformImportResult(
            platform=platform,
            package_kind="excel",
            candidates=results,
            imported=imported,
            rejected=rejected,
        )

    async def _persist_one(
        self,
        *,
        team_id: uuid.UUID,
        platform: Platform,
        mapped: dict[str, Any],
        raw: dict[str, str],
    ) -> ImportedCandidateItem:
        """把单行映射结果落库（结构化路径，跳过 parser）。"""
        # 归一化字段值
        if mapped.get("education"):
            mapped["education"] = _EDU_MAP.get(
                str(mapped["education"]).strip().lower(), "other"
            )
        if mapped.get("gender"):
            mapped["gender"] = _GENDER_MAP.get(
                str(mapped["gender"]).strip().lower(), "unknown"
            )
        for int_f in ("age", "years_experience"):
            if mapped.get(int_f):
                try:
                    mapped[int_f] = int(re.sub(r"\D", "", str(mapped[int_f])) or 0)
                except ValueError:
                    mapped[int_f] = None

        # 校验：必须有 name + (phone 或 email)
        name = (mapped.get("name") or "").strip()
        phone = (mapped.get("phone") or "").strip() or None
        email = (mapped.get("email") or "").strip() or None
        if not name:
            return ImportedCandidateItem(
                name="(unknown)",
                status="rejected",
                reject_reason="invalid_structure",
            )
        if not phone and not email:
            return ImportedCandidateItem(
                name=name,
                status="rejected",
                reject_reason="missing_identity",
            )

        try:
            struct = CandidateStructure(
                name=name,
                phone=phone,
                email=email,
                gender=mapped.get("gender"),
                age=mapped.get("age"),
                education=mapped.get("education"),
                years_experience=mapped.get("years_experience"),
                applied_position=mapped.get("applied_position"),
                current_company=mapped.get("current_company"),
                current_title=mapped.get("current_title"),
                location=mapped.get("location"),
                raw=raw,
            )
        except Exception as exc:  # noqa: BLE001
            logger.warning(
                "platform_structure_invalid",
                platform=platform,
                error=str(exc),
            )
            return ImportedCandidateItem(
                name=name,
                status="rejected",
                reject_reason="invalid_structure",
            )

        # TODO(task-14): 真实 dedup_key 由 dedup service 接管
        identity = _normalize_identity(
            {"name": struct.name, "phone": struct.phone, "email": struct.email}
        )
        if identity is None:
            return ImportedCandidateItem(
                name=name, status="rejected", reject_reason="missing_identity"
            )
        dedup_key = f"platform:{platform}:{identity}"

        existing = await self.db.scalar(
            select(Candidate).where(Candidate.dedup_key == dedup_key)
        )
        if existing is not None:
            return ImportedCandidateItem(
                candidate_id=existing.id,
                name=struct.name,
                status="rejected",
                reject_reason="duplicate",
            )

        candidate = Candidate(
            team_id=team_id,
            dedup_key=dedup_key,
            name=struct.name,
            phone=struct.phone,
            email=struct.email,
        )
        self.db.add(candidate)
        await self.db.flush()

        source = CandidateSource(
            candidate_id=candidate.id,
            source_type="platform",
            source_meta={"platform": platform, "kind": "structured"},
        )
        self.db.add(source)
        await self.db.flush()

        # 结构化数据无需 file_storage_key，但仍需一条 candidate_resumes 承载 parsed_structure
        # 用占位 file_storage_key 标记 "no_file"（task 13 不会处理此 resume）
        resume = CandidateResume(
            candidate_id=candidate.id,
            source_id=source.id,
            file_storage_key=f"platform:{platform}:{candidate.id}",
            file_mime=None,
            parsed_text=None,
            parse_status="success",  # 跳过 parser
        )
        self.db.add(resume)
        await self.db.flush()

        # 直接写 ParsedStructure.data（结构化数据已就绪，无需 LLM extractor）
        # 延迟 import 避免循环
        from app.models.candidate import ParsedStructure

        self.db.add(
            ParsedStructure(
                resume_id=resume.id,
                data=struct.model_dump(mode="json"),
            )
        )
        await self.db.flush()

        return ImportedCandidateItem(
            candidate_id=candidate.id,
            resume_id=resume.id,
            name=struct.name,
            status="ok",
        )

    # ----- 附件包 ZIP -----

    async def _import_attachment_zip(
        self,
        *,
        team_id: uuid.UUID,
        platform: Platform,
        content: bytes,
    ) -> PlatformImportResult:
        """附件包：解压每个简历 → MIME 嗅探 → storage.put → 写库 + 入队。

        复用任务 9 校验规则（MIME 白名单 + 扩展名一致性）。
        """
        results: list[ImportedCandidateItem] = []
        try:
            with zipfile.ZipFile(io.BytesIO(content)) as zf:
                for info in zf.infolist():
                    if info.is_dir():
                        continue
                    ext = "." + info.filename.rsplit(".", 1)[-1].lower() if "." in info.filename else ""
                    if ext not in _ATTACHMENT_EXT:
                        continue
                    member_bytes = zf.read(info)
                    item = await self._persist_attachment(
                        team_id=team_id,
                        platform=platform,
                        original_name=info.filename,
                        content=member_bytes,
                    )
                    results.append(item)
        except zipfile.BadZipFile as exc:
            raise UnsupportedPlatformError(
                "无法解析 ZIP 包",
                detection=DetectionResult(
                    platform=None,
                    confidence=0.0,
                    package_kind=None,
                    threshold=self.threshold,
                    signals=[],
                    scores={p: 0.0 for p in _FILENAME_KEYWORDS},
                ),
                support_feedback_url=self.feedback_url,
            ) from exc

        imported = sum(1 for r in results if r.status == "ok")
        rejected = len(results) - imported
        return PlatformImportResult(
            platform=platform,
            package_kind="attachment_zip",
            candidates=results,
            imported=imported,
            rejected=rejected,
        )

    async def _persist_attachment(
        self,
        *,
        team_id: uuid.UUID,
        platform: Platform,
        original_name: str,
        content: bytes,
    ) -> ImportedCandidateItem:
        """单个附件落库（仿 file_upload.confirm，但不走签名 URL）。"""
        real_mime = _sniff_mime(content)
        if real_mime not in _PLATFORM_ALLOWED_MIME:
            logger.warning(
                "platform_attachment_mime_rejected",
                platform=platform,
                filename=original_name,
                real_mime=real_mime,
            )
            return ImportedCandidateItem(
                name=original_name,
                status="rejected",
                reject_reason="invalid_structure",
            )

        ext = original_name.rsplit(".", 1)[-1].lower() if "." in original_name else ""
        # 扩展名 ↔ MIME 一致性（同任务 9 规则）
        expected = {
            "pdf": "application/pdf",
            "doc": "application/msword",
            "docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            "png": "image/png",
            "jpg": "image/jpeg",
            "jpeg": "image/jpeg",
        }.get(ext)
        if expected and real_mime != expected:
            return ImportedCandidateItem(
                name=original_name,
                status="rejected",
                reject_reason="invalid_structure",
            )

        file_key = f"{team_id}/{uuid.uuid4()}/{uuid.uuid4()}.{ext or 'bin'}"
        try:
            await self.storage.put(
                file_key, content, mime=real_mime, encrypt=True
            )
        except Exception:  # noqa: BLE001
            logger.exception(
                "platform_attachment_storage_failed",
                filename=original_name,
            )
            return ImportedCandidateItem(
                name=original_name,
                status="rejected",
                reject_reason="storage_error",
            )

        # TODO(task-14): 真实 dedup_key 由 dedup service 接管
        dedup_key = f"platform:{platform}:upload:{file_key}"
        candidate = Candidate(
            team_id=team_id,
            dedup_key=dedup_key,
            name=original_name,
            phone=None,
            email=None,
        )
        self.db.add(candidate)
        await self.db.flush()

        source = CandidateSource(
            candidate_id=candidate.id,
            source_type="platform",
            source_meta={
                "platform": platform,
                "kind": "attachment",
                "original_name": original_name,
            },
        )
        self.db.add(source)
        await self.db.flush()

        resume = CandidateResume(
            candidate_id=candidate.id,
            source_id=source.id,
            file_storage_key=file_key,
            file_mime=real_mime,
            parse_status="pending",
        )
        self.db.add(resume)
        await self.db.flush()

        # 入 async_jobs（task 12 消费 → task 13 parser）
        idem = f"parse:{resume.id}"
        existing_job = await self.db.scalar(
            select(AsyncJob).where(AsyncJob.idempotency_key == idem)
        )
        if existing_job is None:
            self.db.add(
                AsyncJob(
                    task_type="parse",
                    target_id=resume.id,
                    status="queued",
                    idempotency_key=idem,
                    payload={
                        "file_key": file_key,
                        "mime": real_mime,
                        "source": "platform",
                        "platform": platform,
                    },
                )
            )
            await self.db.flush()

        return ImportedCandidateItem(
            candidate_id=candidate.id,
            resume_id=resume.id,
            name=original_name,
            status="ok",
        )


__all__ = ["PlatformImportAdapter", "UnsupportedPlatformError"]
