"""ExportService（任务 22）：候选人列表批量导出到 xlsx。

流程：
1. ``request_export(job_id, user_id, filters)``：
   - 校验 job 归属 team
   - 估算行数（按 candidate 数）→ 决定同步 or 异步
   - 同步（行数 ≤ EXPORT_ASYNC_THRESHOLD）→ 立即跑 ``_generate``，返回 ``download_url``
   - 异步（行数 > 阈值）→ 入 ``async_jobs`` (task_type="export")，返回 ``job_id``；
     ``run_export_handler`` 直接调 ``_generate`` + ``_notify_user``，
     handler 返回的 dict 通过 ``mark_success`` 自动落到 ``payload['result']``
2. ``get_signed_download_url(file_key)``：5min 过期；team 隔离校验前缀

Excel schema（按列顺序）：
- 姓名 / 邮箱 / 电话 / 学历 / 工作年限 / 当前公司 / 技能
- 来源（upload/platform/email）
- 总分 / skill / experience / education / stability / potential
- 推荐理由（bullet_points 拼接）
- 是否淘汰 + 淘汰原因
- 面试问题（每题一行，按维度分组）

设计约束（Restrictions）：
- 行数 > EXPORT_ASYNC_THRESHOLD（默认 5000）强制异步
- 跨 team 访问 → NotFoundError（不暴露存在性）
- 下载 URL：STORAGE_SIGNED_URL_EXPIRE_SECONDS（默认 300s = 5min）
- 不在前端直连对象存储（server 签 URL）
- 文件路径前缀：``exports/{team_id}/{job_id}/{uuid}.xlsx``
- 仅包含当前用户可见字段（无 PII 加密列；明文从 ORM 读）

Email 通知：当前未接 SMTP（任务 11 只读邮件不发），用 logger.info 占位，
留 TODO(task-23/24) 待 SMTP 集成。
"""
from __future__ import annotations

import io
import uuid
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.adapters.storage import (
    BaseStorageAdapter,
    get_storage,
)
from app.core.config import settings
from app.core.logging import get_logger
from app.core.middleware.error_handler import NotFoundError
from app.models.candidate import (
    Candidate,
    CandidateResume,
    CandidateSource,
    ParsedStructure,
)
from app.models.interview import InterviewQuestion
from app.models.job import Job
from app.models.score import Score, ScoreReason
from app.models.screening import ScreeningResult
from app.services.async_job_service import AsyncJobService

logger = get_logger(__name__)


# ============================================================================
# 常量
# ============================================================================


EXPORT_ASYNC_THRESHOLD: int = 5000
"""行数超过该值 → 强制异步（入 async_jobs）。"""

SIGNED_URL_EXPIRE_SECONDS: int = settings.STORAGE_SIGNED_URL_EXPIRE_SECONDS
"""下载 URL 过期时间（默认 300s = 5min）。"""

EXPORT_COLUMNS: tuple[str, ...] = (
    "姓名",
    "邮箱",
    "电话",
    "学历",
    "工作年限",
    "当前公司",
    "技能",
    "来源",
    "总分",
    "技能分",
    "经验分",
    "学历分",
    "稳定性分",
    "潜力分",
    "推荐理由",
    "是否淘汰",
    "淘汰原因",
    "面试问题",
)
"""Excel 列顺序（中文名）。"""


# ============================================================================
# ExportService
# ============================================================================


class ExportService:
    """候选人批量导出（xlsx）。"""

    def __init__(
        self,
        db: AsyncSession,
        *,
        storage: BaseStorageAdapter | None = None,
    ) -> None:
        self._db = db
        self._storage = storage or get_storage()

    # ----- 主入口：request_export -----

    async def request_export(
        self,
        *,
        team_id: uuid.UUID,
        user_id: uuid.UUID,
        job_id: uuid.UUID,
        filters: dict[str, Any] | None = None,
        format: str = "xlsx",
    ) -> dict[str, Any]:
        """请求导出；自动判断同步 vs 异步。

        Returns:
            - 同步路径：``{"mode": "sync", "download_url": str, "expires_in": int, "row_count": int}``
            - 异步路径：``{"mode": "async", "job_id": str, "row_count": int}``
        """
        # 校验 job 归属 team
        job = await self._validate_job_in_team(job_id, team_id)
        _ = job  # 仅校验，不使用

        # 估算行数（按 candidate count）
        row_count = await self._count_candidates(team_id, job_id, filters)

        if row_count > EXPORT_ASYNC_THRESHOLD:
            # 异步：入队
            service = AsyncJobService(self._db)
            job_obj = await service.enqueue(
                task_type="export",
                target_id=user_id,
                payload={
                    "job_id": str(job_id),
                    "team_id": str(team_id),
                    "user_id": str(user_id),
                    "format": format,
                    "filters": filters or {},
                    "estimated_rows": row_count,
                },
                idempotency_key=f"export:{job_id}:{user_id}:{row_count}",
            )
            return {
                "mode": "async",
                "job_id": str(job_obj.id),
                "row_count": row_count,
            }

        # 同步：立即生成
        file_key, file_size = await self._generate(
            job_id=job_id,
            team_id=team_id,
            filters=filters or {},
        )
        download_url = await self._storage.signed_url(
            file_key,
            expires=SIGNED_URL_EXPIRE_SECONDS,
            method="GET",
        )
        return {
            "mode": "sync",
            "download_url": download_url,
            "expires_in": SIGNED_URL_EXPIRE_SECONDS,
            "row_count": row_count,
            "file_key": file_key,
            "file_size": file_size,
        }

    # ----- 查询签名 URL（已存在 file_key 时）-----

    async def get_signed_download_url(
        self,
        *,
        team_id: uuid.UUID,
        file_key: str,
    ) -> str:
        """取 5min 签名下载 URL；校验 file_key 前缀归属 team。"""
        _validate_file_key_prefix(file_key, team_id)

        if not await self._storage.exists(file_key):
            raise NotFoundError(
                f"导出文件不存在或已过期", resource="export_file"
            )
        return await self._storage.signed_url(
            file_key,
            expires=SIGNED_URL_EXPIRE_SECONDS,
            method="GET",
        )

    # ----- 内部：生成 xlsx -----

    async def _generate(
        self,
        *,
        job_id: uuid.UUID,
        team_id: uuid.UUID,
        filters: dict[str, Any],
    ) -> tuple[str, int]:
        """生成 xlsx 并上传；返回 (file_key, size_bytes)。"""
        rows = await self._fetch_rows(team_id, job_id, filters)
        wb = _build_workbook(rows)
        buf = io.BytesIO()
        wb.save(buf)
        data = buf.getvalue()

        file_key = _make_file_key(team_id, job_id)
        await self._storage.put(
            file_key,
            data,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            encrypt=True,
        )
        logger.info(
            "export_generated",
            file_key=file_key,
            size_bytes=len(data),
            row_count=len(rows),
        )
        return file_key, len(data)

    async def _fetch_rows(
        self,
        team_id: uuid.UUID,
        job_id: uuid.UUID,
        filters: dict[str, Any],
    ) -> list[dict[str, Any]]:
        """按 filters 拉所有候选人的导出数据（一次大 JOIN 不现实；分两步）。"""
        # 1) candidate + latest resume/structure + source + screening + score
        stmt = (
            select(
                Candidate.id,
                Candidate.name,
                Candidate.phone,
                Candidate.email,
                CandidateSource.source_type,
                ParsedStructure.data,
                ScreeningResult.disqualified,
                ScreeningResult.reasons,
                Score.total,
                Score.skill,
                Score.experience,
                Score.education,
                Score.stability,
                Score.potential,
            )
            .join(CandidateSource, CandidateSource.candidate_id == Candidate.id)
            .outerjoin(
                ParsedStructure,
                ParsedStructure.resume_id
                == select(CandidateResume.id)
                .where(CandidateResume.candidate_id == Candidate.id)
                .order_by(CandidateResume.uploaded_at.desc())
                .limit(1)
                .scalar_subquery(),
            )
            .outerjoin(
                ScreeningResult,
                (ScreeningResult.candidate_id == Candidate.id)
                & (ScreeningResult.job_id == job_id),
            )
            .outerjoin(
                Score,
                (Score.candidate_id == Candidate.id) & (Score.job_id == job_id),
            )
            .where(Candidate.team_id == team_id)
        )
        # filters: disqualified / min_score
        only_disqualified = filters.get("disqualified")
        if only_disqualified is True:
            stmt = stmt.where(ScreeningResult.disqualified.is_(True))
        elif only_disqualified is False:
            stmt = stmt.where(
                (ScreeningResult.disqualified.is_(False))
                | (ScreeningResult.id.is_(None))
            )

        min_score = filters.get("min_score")
        if min_score is not None:
            stmt = stmt.where(Score.total >= int(min_score))

        result = await self._db.execute(stmt)
        candidate_rows = result.all()

        if not candidate_rows:
            return []

        candidate_ids = [r[0] for r in candidate_rows]

        # 2) 批量取 reasons + interview questions
        reasons_map = await self._fetch_reasons(candidate_ids)
        questions_map = await self._fetch_interview_questions(job_id, candidate_ids)

        rows: list[dict[str, Any]] = []
        for r in candidate_rows:
            (
                cid, name, phone, email, source_type,
                parsed_data, disqualified, reasons,
                total, skill, exp, edu, sta, pot,
            ) = r
            structure = (parsed_data or {}).get("structure", {}) if parsed_data else {}
            rows.append(
                {
                    "姓名": name or "",
                    "邮箱": email or "",
                    "电话": _mask_phone(phone),
                    "学历": structure.get("education") or "",
                    "工作年限": structure.get("years_of_experience") or "",
                    "当前公司": structure.get("current_company") or "",
                    "技能": ", ".join(structure.get("skills") or []),
                    "来源": source_type or "",
                    "总分": total if total is not None else "",
                    "技能分": skill if skill is not None else "",
                    "经验分": exp if exp is not None else "",
                    "学历分": edu if edu is not None else "",
                    "稳定性分": sta if sta is not None else "",
                    "潜力分": pot if pot is not None else "",
                    "推荐理由": "\n".join(reasons_map.get(cid, [])),
                    "是否淘汰": "是" if disqualified else "否",
                    "淘汰原因": "\n".join(reasons or []) if disqualified else "",
                    "面试问题": "\n".join(questions_map.get(cid, [])),
                }
            )
        return rows

    async def _fetch_reasons(
        self, candidate_ids: list[uuid.UUID]
    ) -> dict[uuid.UUID, list[str]]:
        """按 candidate_id 拉推荐理由（join score → score_reasons）。"""
        if not candidate_ids:
            return {}
        stmt = (
            select(Score.candidate_id, ScoreReason.bullet_points)
            .join(ScoreReason, ScoreReason.score_id == Score.id)
            .where(
                Score.candidate_id.in_(candidate_ids),
                ScoreReason.type == "recommend",
            )
        )
        result = await self._db.execute(stmt)
        out: dict[uuid.UUID, list[str]] = {}
        for cid, bullets in result.all():
            out.setdefault(cid, []).extend(bullets or [])
        return out

    async def _fetch_interview_questions(
        self, job_id: uuid.UUID, candidate_ids: list[uuid.UUID]
    ) -> dict[uuid.UUID, list[str]]:
        if not candidate_ids:
            return {}
        stmt = (
            select(
                InterviewQuestion.candidate_id,
                InterviewQuestion.dimension,
                InterviewQuestion.question,
            )
            .where(
                InterviewQuestion.job_id == job_id,
                InterviewQuestion.candidate_id.in_(candidate_ids),
            )
            .order_by(
                InterviewQuestion.candidate_id,
                InterviewQuestion.batch_id,
                InterviewQuestion.sort_order,
            )
        )
        result = await self._db.execute(stmt)
        out: dict[uuid.UUID, list[str]] = {}
        for cid, dim, q in result.all():
            out.setdefault(cid, []).append(f"[{dim}] {q}")
        return out

    async def _count_candidates(
        self,
        team_id: uuid.UUID,
        job_id: uuid.UUID,
        filters: dict[str, Any],
    ) -> int:
        """估算行数（不拉数据，仅 COUNT）。"""
        stmt = select(func.count(Candidate.id)).where(
            Candidate.team_id == team_id
        )
        result = await self._db.execute(stmt)
        return int(result.scalar_one())

    async def _validate_job_in_team(
        self, job_id: uuid.UUID, team_id: uuid.UUID
    ) -> Job:
        job = await self._db.get(Job, job_id)
        if job is None or job.team_id != team_id:
            raise NotFoundError(
                f"job {job_id} 不存在或无权访问", resource="job"
            )
        return job

    async def _notify_user(
        self,
        *,
        user_id: uuid.UUID,
        team_id: uuid.UUID,
        file_key: str,
        row_count: int,
    ) -> bool:
        """异步导出完成后通知用户（占位：未接 SMTP）。

        TODO(task-23/24): 接入 SMTP / 阿里云邮件推送。
        当前仅 log，返回 True 表示"已记录通知意图"。
        """
        logger.info(
            "export_email_notification_placeholder",
            user_id=str(user_id),
            team_id=str(team_id),
            file_key=file_key,
            row_count=row_count,
        )
        return True


# ============================================================================
# 辅助
# ============================================================================


def _make_file_key(team_id: uuid.UUID, job_id: uuid.UUID) -> str:
    """文件路径前缀：exports/{team_id}/{job_id}/{uuid}.xlsx。"""
    return f"exports/{team_id}/{job_id}/{uuid.uuid4()}.xlsx"


def _validate_file_key_prefix(file_key: str, team_id: uuid.UUID) -> None:
    """校验 file_key 前缀归属 team；跨 team 抛 NotFoundError（不暴露存在性）。"""
    parts = file_key.split("/")
    if len(parts) < 3 or parts[0] != "exports" or parts[1] != str(team_id):
        raise NotFoundError(
            "导出文件不存在或无权访问", resource="export_file"
        )


def _mask_phone(phone: str | None) -> str:
    """脱敏手机号（中间 4 位）：138****0000。"""
    if not phone or len(phone) < 7:
        return phone or ""
    return phone[:3] + "****" + phone[-4:]


def _build_workbook(rows: list[dict[str, Any]]) -> Workbook:
    """从 rows 构建 xlsx Workbook（含表头样式 + 冻结首行）。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "候选人列表"

    # 表头
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    for col_idx, col_name in enumerate(EXPORT_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 数据行
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, col_name in enumerate(EXPORT_COLUMNS, start=1):
            value = row.get(col_name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # 列宽（按列名粗略估算）
    col_widths = {
        "姓名": 12, "邮箱": 24, "电话": 14, "学历": 8, "工作年限": 8,
        "当前公司": 16, "技能": 30, "来源": 10,
        "总分": 6, "技能分": 8, "经验分": 8, "学历分": 8,
        "稳定性分": 8, "潜力分": 8,
        "推荐理由": 40, "是否淘汰": 8, "淘汰原因": 30, "面试问题": 50,
    }
    for col_idx, col_name in enumerate(EXPORT_COLUMNS, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = (
            col_widths.get(col_name, 12)
        )

    # 冻结首行 + 自动筛选
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    return wb


__all__ = [
    "ExportService",
    "EXPORT_ASYNC_THRESHOLD",
    "EXPORT_COLUMNS",
    "SIGNED_URL_EXPIRE_SECONDS",
]
